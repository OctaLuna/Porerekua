#!/usr/bin/env python3
"""
Genera scripts/migrate-geography.sql leyendo el Excel oficial de datos geográficos.
Fuente: docs/AMAZONÍA MAPEO OFICIAL PROYECTO AMAZONIA (1).xlsx
  Columna A = departamento | Columna C = municipio | Columna D = comunidad indígena

Uso: python scripts/generate-geography-sql.py
"""
import sys
from pathlib import Path
from datetime import date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
OUTPUT_PATH = SCRIPT_DIR / "migrate-geography.sql"

# Buscar el Excel por extensión (el nombre usa forma Unicode NFD en el filesystem)
_xlsx = list((PROJECT_ROOT / "docs").glob("*.xlsx"))
if not _xlsx:
    print("ERROR: No se encontró ningún .xlsx en docs/")
    sys.exit(1)
EXCEL_PATH = _xlsx[0]

# Los 9 departamentos de Bolivia con IDs fijos
DEPARTAMENTOS = [
    (1, "Pando",      True),
    (2, "Beni",       True),
    (3, "La Paz",     True),
    (4, "Cochabamba", True),
    (5, "Santa Cruz", True),
    (6, "Oruro",      False),
    (7, "Potosí",     False),
    (8, "Chuquisaca", False),
    (9, "Tarija",     False),
]
DEPT_NAME_TO_ID = {n: i for i, n, _ in DEPARTAMENTOS}

# Correcciones de nombres de municipios conocidos (Excel → nombre correcto)
MUNI_FIXES: dict[str, str] = {
    "San Julian":              "San Julián",
    "San Juan de yapacani":    "San Juan de Yapacaní",
    "Entre Rios (Bulo Bulo )": "Entre Ríos (Bulo Bulo)",
    "Entre Ríos (Bulo Bulo )": "Entre Ríos (Bulo Bulo)",
    "Tapacari":                "Tapacarí",
}

# Correcciones de nombres de comunidades (Excel → nombre correcto)
COMU_FIXES: dict[str, str] = {
    "Reyesano/ Moropa":         "Reyesano / Moropa",
    "ESE EJJA":                 "Esse Ejja",  # mismo pueblo, nombre duplicado
    "Consejo Indigena Yuqui Consejo Indigena del Rio Ichilo YUQUI CIRI":
        "Consejo Indígena Yuqui Consejo Indígena del Río Ichilo (YUQUI CIRI)",
    "Trinitario Moxeño-Territorio Indigena del Parque Isiboro Sécure Consejo Indigena del Sur (TIPNIS)":
        "Trinitario Moxeño - Territorio Indígena del Parque Isiboro Sécure Consejo Indígena del Sur (TIPNIS)",
    "Consejo Regional T´smane Mosetenes Pilón Lajas (CRTM-PL)":
        "Consejo Regional T'simane Mosetenes Pilón Lajas (CRTM-PL)",
}


def is_abbreviation_token(s: str) -> bool:
    """True si el token es una abreviación: sin espacios, todo mayúsculas, alfa+guiones."""
    if not s or " " in s:
        return False
    cleaned = s.replace("-", "").replace("/", "")
    return cleaned.isupper() and cleaned.isalpha() and 2 <= len(cleaned) <= 10


def process_community_cell(v) -> str | None:
    """
    Procesa el valor de una celda de comunidad del Excel.
    Maneja:
      - Celdas con salto de línea + abreviación: 'Nombre completo\nABRV' → 'Nombre completo (ABRV)'
      - Celdas con abreviación como última palabra: 'Nombre ABRV' → 'Nombre (ABRV)'
        (solo si la penúltima palabra NO es todo mayúsculas, para evitar falsos positivos)
      - Trailing/leading spaces y newlines espurios
    """
    if v is None:
        return None

    raw = str(v)
    # Separar por newlines
    lines = [l.strip() for l in raw.split("\n")]
    lines = [l for l in lines if l and l.lower() != "none"]

    if not lines:
        return None

    name = lines[0]

    if len(lines) >= 2:
        # Segunda línea no vacía: puede ser abreviación
        abbr = lines[1]
        if is_abbreviation_token(abbr):
            name = f"{name} ({abbr})"
    else:
        # Línea única: comprobar si termina con abreviación como palabra suelta
        words = name.split()
        if len(words) >= 2:
            last = words[-1]
            second_last = words[-2]
            # Aplicar solo si penúltima palabra NO es todo mayúsculas (evita "YUQUI CIRI")
            if is_abbreviation_token(last) and not second_last.isupper():
                name = " ".join(words[:-1]) + f" ({last})"

    return name.strip() or None


def sql_str(s: str) -> str:
    return "'" + s.replace("'", "''") + "'"


def main():
    print(f"Leyendo {EXCEL_PATH.name} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    print(f"  Filas en el libro: {ws.max_row}")

    # --- Parsear filas con carry-forward ---
    # Fila 1: encabezado principal (skipped por min_row=3)
    # Fila 2: encabezado secundario "Departamento | Provincia | Municipio | ..."
    # Filas 3+: datos reales
    raw_rows: list[tuple[str, str, str | None]] = []
    last_dept: str | None = None
    last_muni: str | None = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        dept_raw = process_community_cell(row[0]) if row[0] and str(row[0]).strip() else None
        muni_raw = process_community_cell(row[2]) if row[2] and str(row[2]).strip() else None
        comu_raw = process_community_cell(row[3])

        # dept/muni usan stripped plain string (no community processing needed)
        dept_raw = str(row[0]).strip() if row[0] and str(row[0]).strip().lower() not in ('none', '') else None
        muni_raw = str(row[2]).strip() if row[2] and str(row[2]).strip().lower() not in ('none', '') else None

        if dept_raw:
            last_dept = dept_raw
        if muni_raw:
            last_muni = muni_raw

        if not last_dept or not last_muni:
            continue

        raw_rows.append((last_dept, last_muni, comu_raw))

    print(f"  Filas con datos (carry-forward): {len(raw_rows)}")
    print(f"  Filas con comunidad: {sum(1 for _, _, c in raw_rows if c)}")

    # --- Municipios únicos ---
    muni_pairs: set[tuple[str, str]] = set()
    for dept, muni, _ in raw_rows:
        if dept in DEPT_NAME_TO_ID:
            muni_norm = MUNI_FIXES.get(muni, muni)
            muni_pairs.add((dept, muni_norm))

    muni_list = sorted(muni_pairs, key=lambda x: (DEPT_NAME_TO_ID[x[0]], x[1]))
    muni_to_id: dict[tuple[str, str], int] = {pair: i for i, pair in enumerate(muni_list, start=1)}
    print(f"\nMunicipios únicos: {len(muni_list)}")
    for dept_name in [n for _, n, _ in DEPARTAMENTOS if _]:
        count = sum(1 for d, _ in muni_list if d == dept_name)
        if count:
            print(f"  {dept_name}: {count}")

    # --- Comunidades únicas ---
    comu_set: set[str] = set()
    for _, _, comu in raw_rows:
        if comu:
            comu_norm = COMU_FIXES.get(comu, comu)
            comu_set.add(comu_norm)

    comu_list = sorted(comu_set)
    comu_to_id: dict[str, int] = {c: i for i, c in enumerate(comu_list, start=1)}
    print(f"\nComunidades indígenas únicas: {len(comu_list)}")

    # --- Relaciones municipio ↔ comunidad ---
    relations: set[tuple[int, int]] = set()
    for dept, muni, comu in raw_rows:
        if comu and dept in DEPT_NAME_TO_ID:
            muni_norm = MUNI_FIXES.get(muni, muni)
            comu_norm = COMU_FIXES.get(comu, comu)
            muni_key = (dept, muni_norm)
            if muni_key in muni_to_id and comu_norm in comu_to_id:
                relations.add((muni_to_id[muni_key], comu_to_id[comu_norm]))

    relations_list = sorted(relations)
    print(f"Relaciones municipio↔comunidad: {len(relations_list)}")

    # --- Generar SQL ---
    L: list[str] = []
    HR = "-- " + "=" * 60

    L += [
        HR,
        "-- MIGRACIÓN: Departamentos, Municipios y Comunidades Indígenas",
        "-- Fuente: docs/AMAZONÍA MAPEO OFICIAL PROYECTO AMAZONIA (1).xlsx",
        f"-- Generado: {date.today().isoformat()}",
        HR,
        "",
        "-- 1. Limpiar datos existentes (orden inverso de dependencias)",
        "TRUNCATE TABLE comunidades_municipios RESTART IDENTITY CASCADE;",
        "TRUNCATE TABLE comunidades_indigenas  RESTART IDENTITY CASCADE;",
        "TRUNCATE TABLE municipios             RESTART IDENTITY CASCADE;",
        "TRUNCATE TABLE departamentos          RESTART IDENTITY CASCADE;",
        "",
    ]

    # Departamentos
    L += [
        "-- 2. Los 9 departamentos de Bolivia",
        "--    TRUE  = presencia amazónica (fuente: Excel)",
        "--    FALSE = sin presencia amazónica (sedes empresas/orgs)",
        "INSERT INTO departamentos (id_departamento, nombre, amazonico) VALUES",
    ]
    L.append(",\n".join(
        f"({i}, {sql_str(n)}, {'TRUE' if am else 'FALSE'})"
        for i, n, am in DEPARTAMENTOS
    ) + ";")
    L += [
        "",
        "SELECT setval('departamentos_id_departamento_seq',",
        "       (SELECT MAX(id_departamento) FROM departamentos));",
        "",
    ]

    # Municipios
    L += [
        f"-- 3. Municipios ({len(muni_list)} municipios amazónicos según Excel)",
        "INSERT INTO municipios (id_municipio, nombre, id_departamento) VALUES",
    ]
    L.append(",\n".join(
        f"({muni_to_id[(dept, muni)]}, {sql_str(muni)}, {DEPT_NAME_TO_ID[dept]})"
        for dept, muni in muni_list
    ) + ";")
    L += [
        "",
        "SELECT setval('municipios_id_municipio_seq',",
        "       (SELECT MAX(id_municipio) FROM municipios));",
        "",
    ]

    # Comunidades
    L += [
        f"-- 4. Comunidades indígenas ({len(comu_list)} comunidades únicas)",
        "INSERT INTO comunidades_indigenas (id_comunidad, nombre) VALUES",
    ]
    L.append(",\n".join(
        f"({comu_to_id[c]}, {sql_str(c)})"
        for c in comu_list
    ) + ";")
    L += [
        "",
        "SELECT setval('comunidades_indigenas_id_comunidad_seq',",
        "       (SELECT MAX(id_comunidad) FROM comunidades_indigenas));",
        "",
    ]

    # Relaciones
    L += [
        f"-- 5. Relaciones municipio ↔ comunidad ({len(relations_list)} relaciones)",
        "INSERT INTO comunidades_municipios (id_municipio, id_comunidad) VALUES",
    ]
    L.append(",\n".join(f"({m}, {c})" for m, c in relations_list) + ";")
    L.append("")

    # Verificación
    L += [
        HR,
        "-- VERIFICACIÓN (ejecutar manualmente para confirmar resultados)",
        HR,
        "",
        "-- Distribución amazónico / no amazónico:",
        "-- SELECT amazonico, COUNT(*), STRING_AGG(nombre, ', ' ORDER BY nombre)",
        "-- FROM departamentos GROUP BY amazonico;",
        "-- Esperado: TRUE = 5, FALSE = 4",
        "",
        "-- Totales:",
        "-- SELECT",
        "--   (SELECT COUNT(*) FROM departamentos)                       AS total_departamentos,",
        "--   (SELECT COUNT(*) FROM departamentos WHERE amazonico = true) AS amazonicos,",
        f"--   (SELECT COUNT(*) FROM municipios)                         AS total_municipios,   -- esperado: {len(muni_list)}",
        f"--   (SELECT COUNT(*) FROM comunidades_indigenas)              AS total_comunidades,  -- esperado: {len(comu_list)}",
        f"--   (SELECT COUNT(*) FROM comunidades_municipios)             AS total_relaciones;   -- esperado: {len(relations_list)}",
        "",
        "-- Integridad (ambos deben dar 0):",
        "-- SELECT COUNT(*) FROM municipios WHERE id_departamento IS NULL;",
        "-- SELECT COUNT(*) FROM comunidades_municipios cm",
        "--   LEFT JOIN municipios m ON m.id_municipio = cm.id_municipio",
        "--   WHERE m.id_municipio IS NULL;",
        "",
        "-- Municipios por departamento amazónico:",
        "-- SELECT d.nombre, COUNT(m.id_municipio) AS municipios",
        "-- FROM departamentos d LEFT JOIN municipios m ON m.id_departamento = d.id_departamento",
        "-- WHERE d.amazonico = true GROUP BY d.nombre ORDER BY d.nombre;",
    ]

    sql = "\n".join(L) + "\n"
    OUTPUT_PATH.write_text(sql, encoding="utf-8")
    print(f"\n✓ SQL generado: {OUTPUT_PATH}")
    print(f"  Líneas: {sql.count(chr(10))}")


if __name__ == "__main__":
    main()
